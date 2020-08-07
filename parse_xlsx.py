# -*- coding: utf-8 -*-
# author: kuangdd
# date: 2020/8/7
"""
parse_xlsx
"""
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(Path(__name__).stem)

import os
import re
import json
import shutil
import collections as clt
from functools import partial
from multiprocessing.pool import Pool

import numpy as np
from tqdm import tqdm
from matplotlib import pyplot as plt

import openpyxl
import datetime


def parse_lottery_ticket(xlsx_path="data/lottery_ticket.xlsx", lottery_type="ssq", number_of_numbers=7,
                         start_and_end_date=("1000-01-01", "9000-12-31")):
    workbook = openpyxl.load_workbook(xlsx_path)
    shenames = workbook.sheetnames
    if lottery_type not in shenames:
        print(shenames)  # ['各省市', '测试表']
        return
    worksheet = workbook[lottery_type]
    rows = worksheet.max_row

    date1 = datetime.datetime.strptime(start_and_end_date[0], '%Y-%m-%d').date()
    date2 = datetime.datetime.strptime(start_and_end_date[1], '%Y-%m-%d').date()

    date_lst = list(worksheet.columns)[1][2:rows]
    start_idx, end_idx = -1, -1
    for num, cell in enumerate(date_lst):
        if cell.value is None:
            end_idx = num + 1
            break
        date_cur = datetime.datetime.strptime(cell.value, '%Y-%m-%d').date()
        if date_cur >= date1 and start_idx == -1:
            start_idx = num + 2
        if date_cur >= date2 and end_idx == -1:
            end_idx = num + 2
            break

    data = []
    for rows in list(worksheet.rows)[start_idx:end_idx + 1]:
        tmp = []
        for cell in rows[2:2 + number_of_numbers]:
            tmp.append(cell.value)
        data.append(tmp)

    return data


def match_history(data, betting):
    outdt = clt.defaultdict(int)
    for num in range(len(betting) + 1):
        outdt[num] = 0

    for line in data:
        zhong = set(line) & set(betting)
        outdt[len(zhong)] += 1
    return outdt


def run_random_betting():
    data = parse_lottery_ticket()
    print('彩票开奖次数：{}'.format(len(data)))

    for _ in range(10):
        reds = np.random.choice(list(range(1, 34)), 6, replace=False)
        blues = np.random.choice(list(range(1, 27)), 1, replace=False)
        betting = [*reds, *blues]
        print('投注彩票号码：{}'.format(betting))
        outdt = match_history(data, betting)
        print('中奖号码个数统计：{}'.format(dict(outdt)))
        print('-' * 50)


def hi_xlsx():
    # 获取 工作簿对象
    workbook = openpyxl.load_workbook("data/lottery_ticket.xlsx")
    # 与xlrd 模块的区别
    # wokrbook=xlrd.open_workbook(""DataSource\Economics.xls)

    # 获取工作簿 workbook的所有工作表
    # 在xlrd模块中为 sheetnames=workbook.sheet_names()

    shenames = workbook.sheetnames
    print(shenames)  # ['各省市', '测试表']

    # 获得工作簿的表名后，就可以获得表对象
    worksheet = workbook["ssq"]
    print(worksheet)  # <Worksheet "各省市">

    # 还可以通过如下写法获得表对象
    worksheet1 = workbook[shenames[0]]
    print(worksheet1)  # <Worksheet "测试表">

    # 还可以通过索引方式获取表对象
    worksheet = workbook.worksheets[0]
    print(worksheet)  # <Worksheet "各省市">

    # 经过上述操作，我们已经获得了第一个“表”的“表对象“，接下来可以对表对象进行操作

    name = worksheet.title  # 获取表名
    print(name)  # 各省市
    # 在xlrd中为worksheet.name

    # 获取该表相应的行数和列数
    rows = worksheet.max_row
    columns = worksheet.max_column
    print(rows, columns)  # 32 13
    # 在xlrd中为 worksheet.nrows  worksheet.ncols

    for row in worksheet.rows:
        for cell in row:
            pass
            # print(cell.value, end=" ")
        # print()
    """
    各省市 工资性收入 家庭经营纯收入 财产性收入 转移性收入 食品 衣着 居住 家庭设备及服务 ……
    北京市 5047.4 1957.1 678.8 592.2 1879.0 451.6 859.4 303.5 698.1 844.1 575.8 113.1 ……
    天津市 3247.9 2707.4 126.4 146.3 1212.6 265.3 664.4 122.4 441.3 315.6 263.2 56.1 ……
    """

    for col in worksheet.columns:
        for cell in col:
            pass
            # print(cell.value, end=" ")
        # print()

    '''
    各省市 北京市 天津市 河北省 山西省 内蒙古自治区 辽宁省 吉林省 黑龙江省 上海市 江苏省 浙江省 ……
    工资性收入 5047.4 3247.9 1514.7 1374.3 590.7 1499.5 605.1 654.9 6686.0 3104.8 3575.1 ……
    家庭经营纯收入 1957.1 2707.4 2039.6 1622.9 2406.2 2210.8 2556.7 2521.5 767.7 2271.4  ……
    '''

    for rows in list(worksheet.rows)[1:5]:
        for cell in rows[1:9]:
            print(cell.value, end="\t")
        print()
    '''
    各省市 工资性收入 家庭经营纯收入 
    北京市 5047.4 1957.1 
    天津市 3247.9 2707.4 
    '''

    for i in range(2, 6):
        for j in range(2, 10):
            print(worksheet.cell(row=i, column=j).value, end="\t")
        print()
    '''
    各省市 工资性收入 家庭经营纯收入 
    北京市 5047.4 1957.1 
    天津市 3247.9 2707.4 
    '''


if __name__ == '__main__':
    print(__file__)
    run_random_betting()